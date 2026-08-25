# -*- coding: utf-8 -*-
"""Update the SEO tracker sheet in Google Drive from the SEO output folders.
Writes to EVERY '<L>:\\My Drive' mount found (two Drive accounts may be mounted,
e.g. F: and G:) and first MIRRORS the four output folders between mounts (union,
newer file wins) so no day is stranded on a single account.
Run: python export_seo.py
"""
import os, sys, glob, re, shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Windows Japanese console defaults to cp932 and crashes on unicode — force UTF-8
try:
    sys.stdout.reconfigure(encoding="utf-8"); sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

APP = "AnantaSutra-SEO"
SUBS = ("articles", "audits", "keyword-research", "geo-reports")
BASE = os.path.dirname(os.path.abspath(__file__))
NAVY = "1F3A5F"

def drive_roots():
    """Google Drive is F: ONLY (user-confirmed 2026-07-23). G: is NOT Drive — never write there."""
    if os.path.isdir("F:\\My Drive"):
        return ["F:\\My Drive\\%s" % APP]
    print("ERROR: F:\\My Drive not mounted — Google Drive missing, export skipped")
    return []

def ensure_dirs(root):
    for sub in SUBS:
        os.makedirs(os.path.join(root, sub), exist_ok=True)

def copy_newer(src, dst):
    if not os.path.isfile(src):
        return False
    if (not os.path.exists(dst)) or os.path.getmtime(src) > os.path.getmtime(dst) + 1:
        try:
            shutil.copy2(src, dst); return True
        except (PermissionError, OSError):
            return False
    return False

def sync_staging(roots):
    """Copy anything the daily job wrote into local staging/ onto every Drive mount."""
    staging = os.path.join(BASE, "staging")
    n = 0
    if not os.path.isdir(staging):
        return n
    for sub in SUBS:
        s = os.path.join(staging, sub)
        if not os.path.isdir(s):
            continue
        for f in os.listdir(s):
            for root in roots:
                if copy_newer(os.path.join(s, f), os.path.join(root, sub, f)):
                    n += 1
    return n

def mirror_roots(roots):
    """Two-way union of output folders across mounts — no day stranded on one account."""
    n = 0
    if len(roots) < 2:
        return n
    for sub in SUBS:
        for src_root in roots:
            s = os.path.join(src_root, sub)
            if not os.path.isdir(s):
                continue
            for f in os.listdir(s):
                for dst_root in roots:
                    if dst_root == src_root:
                        continue
                    if copy_newer(os.path.join(s, f), os.path.join(dst_root, sub, f)):
                        n += 1
    return n

def first_heading(path):
    try:
        for line in open(path, encoding="utf-8"):
            if line.startswith("# "):
                return line[2:].strip()
    except Exception:
        pass
    return os.path.basename(path)

def collect(root, sub, pat=r"(\d{4}-\d{2}-\d{2})"):
    rows = {}
    for f in glob.glob(os.path.join(root, sub, "*.md")):
        m = re.search(pat, os.path.basename(f))
        if m:
            rows.setdefault(m.group(1), []).append(f)
    return rows

def build_tracker(root):
    tracker = os.path.join(root, "seo-tracker.xlsx")
    arts = collect(root, "articles"); auds = collect(root, "audits")
    serp = collect(root, "keyword-research"); geo = collect(root, "geo-reports")
    all_dates = sorted(set(list(arts) + list(auds) + list(serp) + list(geo)))

    # preserve THIS mount's manual columns (Published?, URL, Notes)
    manual = {}; folder_link = ""
    if os.path.exists(tracker):
        try:
            old = load_workbook(tracker); ws = old.active
            folder_link = ws["B1"].value or ""
            for row in ws.iter_rows(min_row=4, values_only=True):
                if row and row[0]:
                    manual[str(row[0])[:10]] = {"pub": row[6], "url": row[7], "notes": row[8]}
        except Exception as e:
            print("tracker read warn (%s): %s" % (root, e))

    wb = Workbook(); ws = wb.active; ws.title = "SEO + GEO Tracker"
    ws["A1"] = "Folder share link:"; ws["A1"].font = Font(bold=True, size=9)
    ws["B1"] = folder_link
    ws["A2"] = "Daily SEO+GEO output. articles/ = publish-ready blog drafts · audits/ = on-page issues · keyword-research/ = SERP watch · geo-reports/ = AI-answer optimization."
    ws["A2"].font = Font(italic=True, size=8.5, color="5A6B8C")
    headers = ["Date", "Article (target/title)", "Audit", "SERP watch", "GEO report", "Article file", "Published?", "Live URL", "Notes"]
    widths = [11, 46, 8, 10, 10, 34, 11, 34, 30]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=NAVY)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"

    r = 4
    for d in all_dates:
        art = arts.get(d, [""])[0]
        keep = manual.get(d, {})
        vals = [d,
                first_heading(art) if art else "",
                "yes" if d in auds else "",
                "yes" if d in serp else "",
                "yes" if d in geo else "",
                ("articles/" + os.path.basename(art)) if art else "",
                keep.get("pub") or "", keep.get("url") or "", keep.get("notes") or ""]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v); c.font = Font(size=9)
            c.alignment = Alignment(vertical="top", wrap_text=(i in (2, 9)))
        r += 1
    try:
        wb.save(tracker)
        print("SEO TRACKER updated: %d rows -> %s" % (r - 4, root))
    except PermissionError:
        print("LOCKED: seo-tracker.xlsx open on %s — not updated" % root)

def main():
    roots = drive_roots()
    print("Drive mounts:", roots)
    for root in roots:
        ensure_dirs(root)
    print("synced from staging:", sync_staging(roots))
    print("mirrored between mounts:", mirror_roots(roots))
    for root in roots:
        build_tracker(root)

if __name__ == "__main__":
    main()
