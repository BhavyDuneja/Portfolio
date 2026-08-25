# -*- coding: utf-8 -*-
"""Daily Google Search Console pull -> Drive.
No-op (graceful) until seo-pipeline/gsc-token.json exists (created by authorize_gsc.py).
Writes into the live Google Drive (auto-detected):
  <Drive>/AnantaSutra-SEO/gsc-data/YYYY-MM-DD.md   (snapshot)
  <Drive>/AnantaSutra-SEO/gsc-rank.xlsx            (Top Queries / Top Pages / Trend)
"""
import os, sys, json, datetime
try:
    sys.stdout.reconfigure(encoding="utf-8"); sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass
BASE = os.path.dirname(os.path.abspath(__file__))
TOKEN = os.path.join(BASE, "gsc-token.json")

def drive_root(app):
    """Google Drive is F: ONLY (user-confirmed 2026-07-23). G: is NOT Drive — never write there."""
    return "F:\\My Drive\\%s" % app

DRIVE = drive_root("AnantaSutra-SEO")
SCOPES = ["https://www.googleapis.com/auth/webmasters.readonly"]

def log(*a): print("[gsc]", *a)

def main():
    if not os.path.exists(TOKEN):
        log("no gsc-token.json yet — run authorize_gsc.py once. Skipping."); return
    try:
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
    except Exception as e:
        log("libs missing:", e); return

    creds = Credentials.from_authorized_user_file(TOKEN, SCOPES)
    if not creds.valid:
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            open(TOKEN, "w", encoding="utf-8").write(creds.to_json())
        else:
            log("token invalid & not refreshable — re-run authorize_gsc.py"); return

    svc = build("searchconsole", "v1", credentials=creds, cache_discovery=False)

    # pick the anantasutra property (prefer sc-domain)
    sites = svc.sites().list().execute().get("siteEntry", [])
    cand = [s["siteUrl"] for s in sites if "anantasutra" in s["siteUrl"].lower()
            and s.get("permissionLevel") != "siteUnverifiedUser"]
    if not cand:
        log("no anantasutra property visible to this account. Sites seen:", [s.get("siteUrl") for s in sites]); return
    site = sorted(cand, key=lambda u: (0 if u.startswith("sc-domain") else 1))[0]
    log("property:", site)

    end = datetime.date.today() - datetime.timedelta(days=2)   # GSC data lags ~2 days
    start = end - datetime.timedelta(days=27)
    daterange = {"startDate": start.isoformat(), "endDate": end.isoformat()}

    def q(dims, limit=100):
        body = dict(daterange); body["dimensions"] = dims; body["rowLimit"] = limit
        return svc.searchanalytics().query(siteUrl=site, body=body).execute().get("rows", [])

    queries = q(["query"]); pages = q(["page"]); totals = q([])
    tot = totals[0] if totals else {"clicks": 0, "impressions": 0, "ctr": 0, "position": 0}

    os.makedirs(os.path.join(DRIVE, "gsc-data"), exist_ok=True)
    today = datetime.date.today().isoformat()

    # snapshot md
    md = ["# Google Search Console — %s" % today,
          "_Property: %s · window: %s to %s (GSC lags ~2 days)_\n" % (site, start, end),
          "## Totals (28 days)",
          "- Clicks: **%d**  ·  Impressions: **%d**  ·  CTR: **%.2f%%**  ·  Avg position: **%.1f**\n" % (
              tot.get("clicks", 0), tot.get("impressions", 0), tot.get("ctr", 0) * 100, tot.get("position", 0)),
          "## Top 20 queries (by clicks)",
          "| Query | Clicks | Impr | CTR | Avg pos |",
          "|---|---:|---:|---:|---:|"]
    for r in sorted(queries, key=lambda x: -x.get("clicks", 0))[:20]:
        md.append("| %s | %d | %d | %.1f%% | %.1f |" % (
            r["keys"][0], r.get("clicks", 0), r.get("impressions", 0), r.get("ctr", 0) * 100, r.get("position", 0)))
    open(os.path.join(DRIVE, "gsc-data", today + ".md"), "w", encoding="utf-8").write("\n".join(md))

    # xlsx
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
        path = os.path.join(DRIVE, "gsc-rank.xlsx")
        # preserve trend history
        history = []
        if os.path.exists(path):
            try:
                old = load_workbook(path)
                if "Trend" in old.sheetnames:
                    for row in old["Trend"].iter_rows(min_row=2, values_only=True):
                        if row and row[0] and str(row[0]) != today:
                            history.append(row)
            except Exception as e:
                log("trend read warn:", e)
        wb = Workbook()
        ws = wb.active; ws.title = "Top Queries"
        ws.append(["Query", "Clicks", "Impressions", "CTR %", "Avg Position"])
        for c in ws[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3A5F")
        for r in sorted(queries, key=lambda x: -x.get("clicks", 0)):
            ws.append([r["keys"][0], r.get("clicks", 0), r.get("impressions", 0),
                       round(r.get("ctr", 0) * 100, 2), round(r.get("position", 0), 1)])
        ws2 = wb.create_sheet("Top Pages")
        ws2.append(["Page", "Clicks", "Impressions", "CTR %", "Avg Position"])
        for c in ws2[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3A5F")
        for r in sorted(pages, key=lambda x: -x.get("clicks", 0)):
            ws2.append([r["keys"][0], r.get("clicks", 0), r.get("impressions", 0),
                        round(r.get("ctr", 0) * 100, 2), round(r.get("position", 0), 1)])
        ws3 = wb.create_sheet("Trend")
        ws3.append(["Date", "Clicks", "Impressions", "CTR %", "Avg Position"])
        for c in ws3[1]: c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="1F3A5F")
        for row in history: ws3.append(list(row))
        ws3.append([today, tot.get("clicks", 0), tot.get("impressions", 0),
                    round(tot.get("ctr", 0) * 100, 2), round(tot.get("position", 0), 1)])
        wb.save(path)
        log("wrote gsc-rank.xlsx (%d queries, %d pages)" % (len(queries), len(pages)))
    except PermissionError:
        log("gsc-rank.xlsx open — skipped xlsx")

    log("DONE — clicks=%d impressions=%d avgpos=%.1f" % (
        tot.get("clicks", 0), tot.get("impressions", 0), tot.get("position", 0)))

if __name__ == "__main__":
    main()
